import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO
import tempfile

st.set_page_config(page_title="Smart Excel Diff", layout="wide")

st.title("🚀 Smart Excel Difference Highlighter")


# ---------------- NORMALIZE VALUES ----------------
def normalize(value):
    if value is None:
        return ""
    if isinstance(value, float):
        return round(value, 6)
    return str(value).strip()


# ---------------- SAFE CELL VALUE ----------------
def get_cell_value(value_ws, formula_ws, r, c):
    """
    Priority:
    1. Calculated value
    2. Formula text
    3. Empty
    """
    val = value_ws.cell(r, c).value
    if val is not None:
        return normalize(val)

    formula = formula_ws.cell(r, c).value
    if formula is not None:
        return normalize(formula)

    return ""


# ---------------- COLORS ----------------
MODIFIED = PatternFill(start_color="FFFF00",
                       end_color="FFFF00",
                       fill_type="solid")

ADDED = PatternFill(start_color="90EE90",
                    end_color="90EE90",
                    fill_type="solid")


# ---------------- FILE UPLOAD ----------------
raw_file = st.file_uploader("Upload RAW Excel", type=["xlsx"])
changed_file = st.file_uploader("Upload UPDATED Excel", type=["xlsx"])


if raw_file and changed_file:

    st.info("Processing Excel comparison...")

    # RAW workbook
    raw_bytes = BytesIO(raw_file.getvalue())

    raw_wb_formula = load_workbook(raw_bytes)
    raw_wb_values = load_workbook(BytesIO(raw_file.getvalue()),
                                  data_only=True)

    # UPDATED workbook (OUTPUT BASE)
    changed_wb_formula = load_workbook(
        BytesIO(changed_file.getvalue())
    )

    changed_wb_values = load_workbook(
        BytesIO(changed_file.getvalue()),
        data_only=True
    )

    # Only compare matching sheets
    common_sheets = (
        set(raw_wb_formula.sheetnames)
        & set(changed_wb_formula.sheetnames)
    )

    if not common_sheets:
        st.error("❌ No matching sheets found.")
        st.stop()

    st.success(f"Comparing Sheets: {', '.join(common_sheets)}")

    summary_data = []

    # ---------------- COMPARISON ----------------
    for sheet in common_sheets:

        st.write(f"🔍 Processing: {sheet}")

        raw_f_ws = raw_wb_formula[sheet]
        raw_v_ws = raw_wb_values[sheet]

        new_f_ws = changed_wb_formula[sheet]
        new_v_ws = changed_wb_values[sheet]

        raw_rows = raw_f_ws.max_row
        new_rows = new_f_ws.max_row
        max_cols = new_f_ws.max_column

        modified_count = 0
        added_rows = 0

        max_rows = max(raw_rows, new_rows)

        for r in range(1, max_rows + 1):

            # ---------- ADDED ROW ----------
            if r > raw_rows:
                added_rows += 1
                for c in range(1, max_cols + 1):
                    new_f_ws.cell(r, c).fill = ADDED
                continue

            if r > new_rows:
                continue

            # ---------- CELL CHECK ----------
            for c in range(1, max_cols + 1):

                old_val = get_cell_value(
                    raw_v_ws,
                    raw_f_ws,
                    r,
                    c
                )

                new_val = get_cell_value(
                    new_v_ws,
                    new_f_ws,
                    r,
                    c
                )

                if old_val != new_val:
                    new_f_ws.cell(r, c).fill = MODIFIED
                    modified_count += 1

        summary_data.append(
            [sheet, modified_count, added_rows]
        )

    # ---------------- SUMMARY SHEET ----------------
    summary_ws = changed_wb_formula.create_sheet("Diff_Summary")

    summary_ws.append(
        ["Sheet Name", "Modified Cells", "Added Rows"]
    )

    for row in summary_data:
        summary_ws.append(row)

    # ---------------- DOWNLOAD ----------------
    with tempfile.NamedTemporaryFile(delete=False,
                                     suffix=".xlsx") as tmp:

        changed_wb_formula.save(tmp.name)

        with open(tmp.name, "rb") as f:
            st.download_button(
                "⬇️ Download Smart Diff Excel",
                f,
                file_name="smart_diff.xlsx"
            )

    st.success("✅ Comparison Complete!")
